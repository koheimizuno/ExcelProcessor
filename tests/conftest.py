import pytest
import io
import base64
import openpyxl

@pytest.fixture
def sample_excel_file():
    """Create a sample Excel file for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add some test data
    ws['A1'] = "Test"
    ws['B1'] = 123
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return base64.b64encode(buffer.read()).decode()