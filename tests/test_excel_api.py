import pytest
import openpyxl
import base64
import io
from datetime import UTC, datetime
from openpyxl.styles import Font, PatternFill
import sys
from pathlib import Path

# Alternative import approach
try:
    from src.schemas.models import Processing, Cell, CellRange, ProcessingTarget, PasteTarget
    from src.excel.operations import xlsx_operation
    from src.excel.utils import apply_styles, get_cell_range
    from src.excel.processor import ExcelProcessor
except ModuleNotFoundError:
    # Add parent directory to path if running tests directly
    sys.path.insert(0, str(Path(__file__).parent.parent))
    from src.schemas.models import Processing, Cell, CellRange, ProcessingTarget, PasteTarget
    from src.excel.operations import xlsx_operation
    from src.excel.utils import apply_styles, get_cell_range
    from src.excel.processor import ExcelProcessor

@pytest.fixture
def sample_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    ws.cell(row=1, column=1, value="Test")
    ws.cell(row=1, column=2, value="Data")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value=2)
    
    # Add styling
    ws['A1'].font = Font(bold=True)
    ws['A1'].fill = PatternFill(patternType="solid", fgColor="FF0000")
    
    return wb

@pytest.fixture
def xlsx_op(sample_workbook):
    return xlsx_operation(sample_workbook)

class TestXlsxOperation:
    def test_copy_cells(self, xlsx_op):
        """Test copying cells with values and styles"""
        sheet = xlsx_op.workbook["Sheet1"]
        
        process = Processing(
            processing_type="copy",
            target=ProcessingTarget(
                cells=CellRange(
                    start_cell=Cell(col_letter="A", row=1),
                    end_cell=Cell(col_letter="B", row=2)
                )
            ),
            paste_target=PasteTarget(
                sheet_name="Sheet1",
                cells=CellRange(
                    start_cell=Cell(col_letter="C", row=1),
                    end_cell=Cell(col_letter="D", row=2)
                )
            )
        )
        
        xlsx_op.copy_cells("Sheet1", process)
        
        sheet = xlsx_op.workbook["Sheet1"]
        
        # Original assertions
        assert sheet['C1'].value == "Test"
        assert sheet['D1'].value == "Data"

    def test_set_cells(self, xlsx_op):
        """Test setting cell values and styles"""
        process = Processing(
            processing_type="set_cells",
            target=ProcessingTarget(
                cells=CellRange(
                    start_cell=Cell(col_letter="A", row=1),
                    end_cell=Cell(col_letter="B", row=2)
                ),
                values=[["New", "Values"], [3, 4]],
                styles={
                    "font": {"bold": True},
                    "fill": {"patternType": "solid", "fgColor": "FF0000"}
                }
            )
        )
        
        xlsx_op.set_cells("Sheet1", process)
        
        sheet = xlsx_op.workbook["Sheet1"]
        # Verify values
        assert sheet['A1'].value == "New"
        assert sheet['B1'].value == "Values"
        assert sheet['A2'].value == 3
        assert sheet['B2'].value == 4
        # Verify styles
        assert sheet['A1'].font.bold is True
        assert sheet['A1'].fill.fgColor.rgb == "00FF0000"
        assert sheet['B1'].font.bold is True
        assert sheet['B1'].fill.fgColor.rgb == "00FF0000"

    def test_copy_sheet(self, xlsx_op):
        """Test copying an entire sheet"""
        process = Processing(processing_type="copy_sheet")
        
        xlsx_op.copy_sheet("Sheet1", process)
        
        # Verify sheet was copied
        assert "Sheet1_copy" in xlsx_op.workbook.sheetnames
        
        # Verify content and styles were copied
        original_sheet = xlsx_op.workbook["Sheet1"]
        copied_sheet = xlsx_op.workbook["Sheet1_copy"]
        
        assert copied_sheet['A1'].value == original_sheet['A1'].value
        assert copied_sheet['A1'].font.__dict__ == original_sheet['A1'].font.__dict__
        assert copied_sheet['A1'].fill.__dict__ == original_sheet['A1'].fill.__dict__

    def test_insert_sheet(self, xlsx_op):
        """Test inserting a new sheet"""
        process = Processing(processing_type="insert_sheet")
        
        xlsx_op.insert_sheet("NewSheet", process)
        
        # Verify new sheet exists
        assert "NewSheet" in xlsx_op.workbook.sheetnames
        # Verify it's empty
        new_sheet = xlsx_op.workbook["NewSheet"]
        assert new_sheet['A1'].value is None

    def test_delete_sheet(self, xlsx_op):
        """Test deleting a sheet"""
        # Create a sheet to delete
        xlsx_op.workbook.create_sheet("ToDelete")
        process = Processing(processing_type="delete_sheet")
        
        xlsx_op.delete_sheet("ToDelete", process)
        
        assert "ToDelete" not in xlsx_op.workbook.sheetnames

    def test_join_cells(self, xlsx_op):
        """Test joining/merging cells"""
        process = Processing(
            processing_type="join_cells",
            target=ProcessingTarget(
                cells=CellRange(
                    start_cell=Cell(col_letter="A", row=1),
                    end_cell=Cell(col_letter="B", row=2)
                )
            )
        )
        
        xlsx_op.join_cells("Sheet1", process)
        
        sheet = xlsx_op.workbook["Sheet1"]
        def is_cell_merged(coord):
            for merged_range in sheet.merged_cells.ranges:
                if coord in merged_range:
                    return True
            return False
    
        # Verify all cells in range are merged
        assert is_cell_merged("A1")
        assert is_cell_merged("B1")
        assert is_cell_merged("A2")
        assert is_cell_merged("B2")
        
        # Verify merged range boundaries
        merged_range_coords = set(str(mr) for mr in sheet.merged_cells.ranges)
        assert "A1:B2" in merged_range_coords
        
        # Verify content of merged range (should contain value from top-left cell)
        assert sheet['A1'].value == "Test"  # Original value from A1
        assert sheet['B1'].value is None    # Other cells in merge range should be empty
        assert sheet['A2'].value is None
        assert sheet['B2'].value is None

class TestExcelStyles:
    @pytest.fixture
    def xlsx_op(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        return xlsx_operation(wb)

    def test_individual_cell_borders(self, xlsx_op):
        """Test applying borders to individual cells"""
        process = Processing(
            processing_type="set_cells",
            target=ProcessingTarget(
                cells=CellRange(
                    start_cell=Cell(col_letter="B", row=2)
                ),
                styles={
                    "border": {
                        "top": {"style": "thin", "color": "FF0000"},
                        "bottom": {"style": "double", "color": "0000FF"},
                        "left": {"style": "medium", "color": "00FF00"},
                        "right": {"style": "thick", "color": "000000"}
                    }
                }
            )
        )
        
        xlsx_op.set_cells("Sheet1", process)
        cell = xlsx_op.workbook["Sheet1"]["B2"]
        
        assert cell.border.top.style == "thin"
        assert cell.border.top.color.rgb == "00FF0000"
        assert cell.border.bottom.style == "double"
        assert cell.border.bottom.color.rgb == "000000FF"
        assert cell.border.left.style == "medium"
        assert cell.border.left.color.rgb == "0000FF00"
        assert cell.border.right.style == "thick"
        assert cell.border.right.color.rgb == "00000000"

    def test_range_outline_border(self, xlsx_op):
        """Test applying border to the outline of a cell range"""
        process = Processing(
            processing_type="set_cells",
            target=ProcessingTarget(
                cells=CellRange(
                    start_cell=Cell(col_letter="B", row=2),
                    end_cell=Cell(col_letter="D", row=4)
                ),
                styles={
                    "border": {
                        "outline": True,
                        "top": {"style": "thick", "color": "000000"},
                        "bottom": {"style": "thick", "color": "000000"},
                        "left": {"style": "thick", "color": "000000"},
                        "right": {"style": "thick", "color": "000000"}
                    }
                }
            )
        )
        
        xlsx_op.set_cells("Sheet1", process)
        sheet = xlsx_op.workbook["Sheet1"]
        
        # Check top border
        for col in range(2, 5):  # B to D
            assert sheet.cell(row=2, column=col).border.top.style == "thick"
        
        # Check bottom border
        for col in range(2, 5):
            assert sheet.cell(row=4, column=col).border.bottom.style == "thick"
        
        # Check left border
        for row in range(2, 5):
            assert sheet.cell(row=row, column=2).border.left.style == "thick"
        
        # Check right border
        for row in range(2, 5):
            assert sheet.cell(row=row, column=4).border.right.style == "thick"

    def test_background_and_font_colors(self, xlsx_op):
        """Test applying background and font colors"""
        process = Processing(
            processing_type="set_cells",
            target=ProcessingTarget(
                cells=CellRange(
                    start_cell=Cell(col_letter="B", row=2)
                ),
                styles={
                    "fill": {
                        "patternType": "solid",
                        "fgColor": "FF0000"  # Red background
                    },
                    "font": {
                        "color": "0000FF",  # Blue text
                        "name": "Arial",
                        "size": 12
                    }
                }
            )
        )
        
        xlsx_op.set_cells("Sheet1", process)
        cell = xlsx_op.workbook["Sheet1"]["B2"]
        
        assert cell.fill.patternType == "solid"
        assert cell.fill.fgColor.rgb == "00FF0000"
        assert cell.font.color.rgb == "000000FF"

    def test_font_properties(self, xlsx_op):
        """Test applying various font properties"""
        process = Processing(
            processing_type="set_cells",
            target=ProcessingTarget(
                cells=CellRange(
                    start_cell=Cell(col_letter="B", row=2)
                ),
                styles={
                    "font": {
                        "name": "Times New Roman",
                        "size": 14,
                        "bold": True,
                        "italic": True,
                        "underline": "single",
                        "strike": True
                    }
                }
            )
        )
        
        xlsx_op.set_cells("Sheet1", process)
        cell = xlsx_op.workbook["Sheet1"]["B2"]
        
        assert cell.font.name == "Times New Roman"
        assert cell.font.size == 14
        assert cell.font.bold is True
        assert cell.font.italic is True
        assert cell.font.underline == "single"
        assert cell.font.strike is True

    def test_row_height_and_column_width(self, xlsx_op):
        """Test setting row height and column width"""
        sheet = xlsx_op.workbook["Sheet1"]
        process = Processing(
            processing_type="set_cells",
            target=ProcessingTarget(
                cells=CellRange(
                    start_cell=Cell(col_letter="B", row=2),
                    end_cell=Cell(col_letter="D", row=4)
                ),
                values=[["Test"]], 
                styles={
                    "row_height": 30,
                    "column_width": 15
                }
            )
        )
        
        xlsx_op.set_cells("Sheet1", process)
        
        # Check row heights
        for row in range(2, 5):
            assert sheet.row_dimensions[row].height == 30
        
        # Check column widths
        for col_letter in ['B', 'C', 'D']:
            assert sheet.column_dimensions[col_letter].width == 15

    def test_multiple_style_combinations(self, xlsx_op):
        """Test applying multiple styles simultaneously"""
        process = Processing(
            processing_type="set_cells",
            target=ProcessingTarget(
                cells=CellRange(
                    start_cell=Cell(col_letter="B", row=2)
                ),
                styles={
                    "font": {
                        "name": "Arial",
                        "size": 12,
                        "bold": True,
                        "color": "FF0000"
                    },
                    "fill": {
                        "patternType": "solid",
                        "fgColor": "FFFF00"
                    },
                    "border": {
                        "top": {"style": "thick", "color": "000000"},
                        "bottom": {"style": "thick", "color": "000000"}
                    },
                    "alignment": {
                        "horizontal": "center",
                        "vertical": "center"
                    }
                }
            )
        )
        
        xlsx_op.set_cells("Sheet1", process)
        cell = xlsx_op.workbook["Sheet1"]["B2"]
        
        # Verify all styles were applied correctly
        assert cell.font.name == "Arial"
        assert cell.font.size == 12
        assert cell.font.bold is True
        assert cell.font.color.rgb == "00FF0000"
        assert cell.fill.patternType == "solid"
        assert cell.fill.fgColor.rgb == "00FFFF00"
        assert cell.border.top.style == "thick"
        assert cell.border.bottom.style == "thick"
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"

class TestUtils:
    def test_apply_styles(self):
        """Test applying styles from one cell to another"""
        wb = openpyxl.Workbook()
        source_cell = wb.active['A1']
        target_cell = wb.active['B1']
        
        # Set source cell styles
        source_cell.font = Font(bold=True, size=12)
        source_cell.fill = PatternFill(patternType="solid", fgColor="FF0000")
        
        apply_styles(source_cell, target_cell)
        
        # Verify styles were copied
        assert target_cell.font.bold == source_cell.font.bold
        assert target_cell.font.size == source_cell.font.size
        assert target_cell.fill.fgColor.rgb == source_cell.fill.fgColor.rgb

    def test_get_cell_range(self):
        """Test getting row and column ranges from CellRange object"""
        cell_range = CellRange(
            start_cell=Cell(col_letter="A", row=1),
            end_cell=Cell(col_letter="C", row=3)
        )
        
        result = get_cell_range(cell_range)
        
        assert result['rows'] == [1, 2, 3]
        assert result['cols'] == [1, 2, 3]  # Column letters converted to numbers

class TestExcelProcessor:
    def test_process_operations(self):
        """Test processing a list of operations"""
        # Create a base64 encoded Excel file
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        wb.save(output)
        excel_data = base64.b64encode(output.getvalue()).decode()
        
        processor = ExcelProcessor(io.BytesIO(base64.b64decode(excel_data)))
        
        operations = [
            Processing(
                processing_type="set_cells",
                target=ProcessingTarget(
                    cells=CellRange(
                        start_cell=Cell(col_letter="A", row=1)
                    ),
                    values=[["Test"]]
                )
            )
        ]
        
        processor.process_operations("Sheet", operations)
        
        assert processor.workbook.active['A1'].value == "Test"

    @pytest.mark.parametrize("processing_type,expected_method", [
        ("copy", "copy_cells"),
        ("set_cells", "set_cells"),
        ("copy_sheet", "copy_sheet"),
        ("insert_sheet", "insert_sheet"),
        ("delete_sheet", "delete_sheet"),
        ("join_cells", "join_cells")
    ])
    def test_get_operation_method(self, processing_type, expected_method):
        """Test getting the correct method for each operation type"""
        # Create a proper Excel file in memory first
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        wb.save(output)
        output.seek(0)  # Reset the buffer position
        
        processor = ExcelProcessor(output)
        
        method = processor._get_operation_method(processing_type)
        
        assert method.__name__ == expected_method

class TestErrorCases:
    def test_invalid_sheet_name(self, xlsx_op):
        """Test handling of invalid sheet name"""
        process = Processing(processing_type="copy_sheet")
        
        with pytest.raises(ValueError, match="Sheet 'NonExistentSheet' not found"):
            xlsx_op.copy_sheet("NonExistentSheet", process)

    def test_missing_target(self, xlsx_op):
        """Test handling of missing target"""
        process = Processing(
            processing_type="set_cells",
            target=None
        )
        
        with pytest.raises(ValueError, match="Target is required for set_cells operation"):
            xlsx_op.set_cells("Sheet1", process)

    def test_duplicate_sheet_name(self, xlsx_op):
        """Test handling of duplicate sheet name"""
        process = Processing(processing_type="insert_sheet")
        
        with pytest.raises(ValueError, match="Sheet Sheet1 already exists"):
            xlsx_op.insert_sheet("Sheet1", process)