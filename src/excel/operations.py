from typing import Optional, Dict, Any, List, Union
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from src.schemas.models import Processing
from src.excel.utils import apply_styles, get_cell_range

class xlsx_operation:
    def __init__(self, workbook: openpyxl.Workbook):
        """Initialize xlsx_operation with a workbook."""
        self.workbook = workbook

    def copy_cells(self, sheet_name: str, process: Processing) -> None:
        """Copy cells from source to target location."""
        if not process.target or not process.paste_target:
            raise ValueError("Both target and paste_target are required for copy operation")

        source_sheet = self.workbook[sheet_name]
        target_sheet = self.workbook[process.paste_target.sheet_name]
        
        source_range = get_cell_range(process.target.cells)
        paste_start = process.paste_target.cells["starting_point"]
        
        for i, row in enumerate(source_range['rows']):
            for j, col in enumerate(source_range['cols']):
                source_cell = source_sheet.cell(row=row, column=col)
                target_row = paste_start.row + i
                target_col = column_index_from_string(paste_start.col_letter) + j
                
                if process.paste_target.is_insert:
                    if i == 0:  # Only insert once per column
                        target_sheet.insert_cols(target_col)
                    if j == 0:  # Only insert once per row
                        target_sheet.insert_rows(target_row)
                
                target_cell = target_sheet.cell(row=target_row, column=target_col)
                target_cell.value = source_cell.value
                apply_styles(source_cell, target_cell)

    def set_cells(self, sheet_name: str, process: Processing) -> None:
        """Set cell values and styles."""
        if not process.target:
            raise ValueError("Target is required for set_cells operation")

        sheet = self.workbook[sheet_name]
        cell_range = get_cell_range(process.target.cells)
        
        # Set values if provided
        if process.target.values:
            for i, row_values in enumerate(process.target.values):
                for j, value in enumerate(row_values):
                    cell = sheet.cell(
                        row=cell_range['rows'][0] + i,
                        column=cell_range['cols'][0] + j
                    )
                    cell.value = value
        
        # Set styles if provided
        if process.target.styles:
            self._apply_styles_to_range(sheet, cell_range, process.target.styles)

    def _apply_styles_to_range(self, sheet, cell_range: Dict[str, List[int]], styles: Dict[str, Any]) -> None:
        """Apply styles to a range of cells."""
        for row in cell_range['rows']:
            for col in cell_range['cols']:
                cell = sheet.cell(row=row, column=col)
                
                # Apply individual cell styles
                if 'cells' in styles:
                    cell_style = styles['cells'].get(f"{get_column_letter(col)}{row}")
                    if cell_style:
                        apply_styles(None, cell, cell_style)
                
                # Apply border styles
                if 'border' in styles:
                    self._apply_border_styles(cell, styles['border'], row, col, cell_range)

    def _apply_border_styles(self, cell, border_styles: Dict[str, Any], row: int, col: int, cell_range: Dict[str, List[int]]) -> None:
        """Apply border styles to a cell based on its position in the range."""
        is_top = row == cell_range['rows'][0]
        is_bottom = row == cell_range['rows'][-1]
        is_left = col == cell_range['cols'][0]
        is_right = col == cell_range['cols'][-1]

        border_sides = {}
        if is_top and border_styles.get('top'):
            border_sides['top'] = Side(**border_styles['top'])
        if is_bottom and border_styles.get('bottom'):
            border_sides['bottom'] = Side(**border_styles['bottom'])
        if is_left and border_styles.get('left'):
            border_sides['left'] = Side(**border_styles['left'])
        if is_right and border_styles.get('right'):
            border_sides['right'] = Side(**border_styles['right'])

        if border_sides:
            cell.border = Border(**border_sides)

    def copy_sheet(self, sheet_name: str, process: Processing) -> None:
        """Copy entire sheet."""
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name} does not exist")
            
        source = self.workbook[sheet_name]
        target = self.workbook.copy_worksheet(source)
        target.title = f"{sheet_name}_copy"

    def copy_style(self, sheet_name: str, process: Processing) -> None:
        """Copy cell styles from source to target location."""
        if not process.target or not process.paste_target:
            raise ValueError("Both target and paste_target are required for copy_style operation")

        source_sheet = self.workbook[sheet_name]
        target_sheet = self.workbook[process.paste_target.sheet_name]
        
        source_range = get_cell_range(process.target.cells)
        paste_start = process.paste_target.cells["starting_point"]
        
        for i, row in enumerate(source_range['rows']):
            for j, col in enumerate(source_range['cols']):
                source_cell = source_sheet.cell(row=row, column=col)
                target_cell = target_sheet.cell(
                    row=paste_start.row + i,
                    column=column_index_from_string(paste_start.col_letter) + j
                )
                apply_styles(source_cell, target_cell)

    def insert_sheet(self, sheet_name: str, process: Processing) -> None:
        """Insert a new sheet."""
        if sheet_name in self.workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name} already exists")
        self.workbook.create_sheet(sheet_name)

    def delete_sheet(self, sheet_name: str, process: Processing) -> None:
        """Delete specified sheet."""
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name} does not exist")
        del self.workbook[sheet_name]

    def insert_rows_or_cols(self, sheet_name: str, process: Processing) -> None:
        """Insert rows or columns."""
        if not process.target or not process.target.cells:
            raise ValueError("Target cells are required for insert operation")

        sheet = self.workbook[sheet_name]
        cell_range = get_cell_range(process.target.cells)
        
        if process.target.cells.start_cell.row and not process.target.cells.start_cell.col_letter:
            # Insert rows
            sheet.insert_rows(cell_range['rows'][0], len(cell_range['rows']))
        else:
            # Insert columns
            sheet.insert_cols(cell_range['cols'][0], len(cell_range['cols']))

    def delete_rows_or_cols(self, sheet_name: str, process: Processing) -> None:
        """Delete rows or columns."""
        if not process.target or not process.target.cells:
            raise ValueError("Target cells are required for delete operation")

        sheet = self.workbook[sheet_name]
        cell_range = get_cell_range(process.target.cells)
        
        if process.target.cells.start_cell.row and not process.target.cells.start_cell.col_letter:
            # Delete rows
            sheet.delete_rows(cell_range['rows'][0], len(cell_range['rows']))
        else:
            # Delete columns
            sheet.delete_cols(cell_range['cols'][0], len(cell_range['cols']))

    def hide_rows_or_cols(self, sheet_name: str, process: Processing) -> None:
        """Hide rows or columns."""
        if not process.target or not process.target.cells:
            raise ValueError("Target cells are required for hide operation")

        sheet = self.workbook[sheet_name]
        cell_range = get_cell_range(process.target.cells)
        
        if process.target.cells.start_cell.row and not process.target.cells.start_cell.col_letter:
            # Hide rows
            for row in cell_range['rows']:
                sheet.row_dimensions[row].hidden = True
        else:
            # Hide columns
            for col in cell_range['cols']:
                sheet.column_dimensions[get_column_letter(col)].hidden = True

    def set_cells(self, sheet_name: str, process: Processing) -> None:
        """Set cell values and styles."""
        if not process.target:
            raise ValueError("Target is required for set_cells operation")

        sheet = self.workbook[sheet_name]
        cell_range = get_cell_range(process.target.cells)
        
        # Set values if provided
        if process.target.values:
            for i, row_values in enumerate(process.target.values):
                for j, value in enumerate(row_values):
                    cell = sheet.cell(
                        row=cell_range['rows'][0] + i,
                        column=cell_range['cols'][0] + j
                    )
                    cell.value = value
        
        # Set styles if provided
        if process.target.styles:
            for row in cell_range['rows']:
                for col in cell_range['cols']:
                    cell = sheet.cell(row=row, column=col)
                    apply_styles(None, cell, process.target.styles)

    def join_cells(self, sheet_name: str, process: Processing) -> None:
        """Merge cells in the specified range."""
        if not process.target or not process.target.cells:
            raise ValueError("Target cells are required for join_cells operation")

        sheet = self.workbook[sheet_name]
        cell_range = get_cell_range(process.target.cells)
        
        start_cell = f"{get_column_letter(cell_range['cols'][0])}{cell_range['rows'][0]}"
        end_cell = f"{get_column_letter(cell_range['cols'][-1])}{cell_range['rows'][-1]}"
        sheet.merge_cells(f"{start_cell}:{end_cell}")