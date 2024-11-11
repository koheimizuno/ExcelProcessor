from typing import Dict, Any, Optional, List
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from src.schemas.models import CellRange

def apply_styles(source_cell: Optional[openpyxl.cell.Cell], 
                target_cell: openpyxl.cell.Cell, 
                styles: Optional[Dict[str, Any]] = None) -> None:
    """Apply styles to a target cell either from a source cell or styles dict."""
    if source_cell:
        # Copy font properties
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
        
        # Copy fill properties
        if source_cell.fill:
            target_cell.fill = PatternFill(
                patternType=source_cell.fill.patternType,
                fgColor=source_cell.fill.fgColor.rgb if source_cell.fill.fgColor else None,
                bgColor=source_cell.fill.bgColor.rgb if source_cell.fill.bgColor else None
            )
        
        # Copy border properties
        if source_cell.border:
            sides = {}
            for side in ['left', 'right', 'top', 'bottom']:
                source_side = getattr(source_cell.border, side)
                if source_side:
                    sides[side] = Side(
                        style=source_side.style,
                        color=source_side.color
                    )
            target_cell.border = Border(**sides)
        
        # Copy alignment
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                textRotation=source_cell.alignment.textRotation,
                wrapText=source_cell.alignment.wrapText,
                shrinkToFit=source_cell.alignment.shrinkToFit,
                indent=source_cell.alignment.indent
            )
        
        # Copy number format as is
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
    
    if styles:
        if 'font' in styles:
            target_cell.font = Font(**styles['font'])
        if 'border' in styles:
            target_cell.border = Border(**styles['border'])
        if 'fill' in styles:
            target_cell.fill = PatternFill(**styles['fill'])
        if 'alignment' in styles:
            target_cell.alignment = Alignment(**styles['alignment'])

def get_cell_range(cell_range: CellRange) -> Dict[str, List[int]]:
    """Convert CellRange to lists of row and column indices."""
    start_col = column_index_from_string(cell_range.start_cell.col_letter) if cell_range.start_cell.col_letter else None
    start_row = cell_range.start_cell.row if cell_range.start_cell.row else None
    
    if cell_range.end_cell:
        end_col = column_index_from_string(cell_range.end_cell.col_letter) if cell_range.end_cell.col_letter else start_col
        end_row = cell_range.end_cell.row if cell_range.end_cell.row else start_row
    else:
        end_col = start_col
        end_row = start_row
    
    return {
        'rows': list(range(start_row, end_row + 1)) if start_row else [],
        'cols': list(range(start_col, end_col + 1)) if start_col else []
    }